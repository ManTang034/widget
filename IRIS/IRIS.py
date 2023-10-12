"""
Created:20230920
LastUpdate:20230928
Context:下载文献
Status:done
Author:Man Tang
"""
import requests
import time
import os
import shutil
import urllib3
import pyautogui
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.options import Options

KeyWords = ['meta-learning', 'rl', 'robustness', 'robotics', 'self-supervision']


class Download_Article(object):
    def __init__(self,
                 download_path,  # 下载路径
                 ):
        self.dp = download_path

    def download(self):
        t = 2
        # 新建一个Excel表格存放信息
        filename = "IRIS_Publications"
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'Reinforcement Learning'
        name = ["Article Title", "Href", "New Href"]

        for a, header in enumerate(name):
            sheet1.cell(row=1, column=a + 1, value=header)

        # 目标url
        url = 'https://irislab.stanford.edu//publications.html'
        chrome_options = Options()
        # chrome_options.add_argument('--headless')  # 无头模式

        # 创建 WebDriver 对象，指明使用chrome浏览器驱动
        wd = webdriver.Chrome(options=chrome_options)
        wd.implicitly_wait(10)

        # 调用WebDriver 对象的get方法 可以让浏览器打开指定网址
        try:
            wd.get(url)
            print('成功访问目标网页')
        except WebDriverException as e:
            print('访问目标网页失败:', e)

        # 分类
        wd.find_element_by_xpath(r'//*[@id="myBtnContainer"]/button[5]').click()

        element = wd.find_element(By.CLASS_NAME, 'row')
        # 获取所有年份
        year_article = element.find_elements(By.CLASS_NAME, 'row')
        # 遍历每个年份,将得到的文章标题与下载链接存放
        result = {}
        for year in year_article:
            # 筛选出每个年份包含rl关键字的所有合规class
            articles = year.find_elements(By.CSS_SELECTOR, ".pubitem[class*='rl']")
            # 遍历每个合规class
            for article in articles:

                title = article.find_element(By.CLASS_NAME, 'pubtitle').text
                link = article.find_element(By.TAG_NAME, 'a').get_attribute('href')
                result[title] = link
                sheet1.cell(row=t, column=1, value=title)
                sheet1.cell(row=t, column=2, value=link)
                title = self.replace_invalid_characters(title)
                if not link.endswith('.pdf'):
                    link = link + '.pdf'
                    link = link.replace('/abs/', '/pdf/')
                    link = link.replace('http:', 'https:')

                # 判断文章年份
                Year = self.which_year(link)

                if Year > 19:
                    # link = link.replace('https://arxiv.org/', 'http://xxx.itp.ac.cn/')
                    sheet1.cell(row=t, column=3, value=link)
                    print(link)
                    wd1 = webdriver.Chrome()
                    wd1.implicitly_wait(10)
                    wd1.get(link)
                    time.sleep(5)
                    title = str(Year) + '-' + title
                    self.simulation_save(title)
                    wd1.quit()
                else:
                    sheet1.cell(row=t, column=3, value=link)
                    print(link)
                    # wd1 = webdriver.Chrome()
                    # wd1.implicitly_wait(10)
                    # wd1.get(link)
                    # time.sleep(25)
                    # title = str(Year) + '-' + title
                    # self.simulation_save(title)
                    # wd1.quit()
                t += 1
        wb.save(filename + '.xlsx')
        wd.quit()

    # 替换标题中不规范的字符
    def replace_invalid_characters(self, articlename):
        # 定义要替换的不规范字符
        invalid_chars = r'[\\/:*?"<>|]'
        replaced_article = re.sub(invalid_chars, '-', articlename)
        return replaced_article

    # 判断当前文章的年份
    def which_year(self, url):
        pattern = r'pdf/(\d{2})(\d{2})'
        match = re.search(pattern, url)

        if match:
            year = int(match.group(1))
        return year
        #     if year > 19:
        #         return year
        #     else:
        #         return False
        # else:
        #     return False

    def simulation_save(self, title):
        print(title)
        pyautogui.hotkey('ctrl', 's')
        time.sleep(1)
        pyautogui.write(title)
        time.sleep(1)
        for i in range(6):
            time.sleep(0.5)
            pyautogui.press('tab')
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.write(self.dp)
        time.sleep(1)
        pyautogui.press('enter')
        for i in range(9):
            time.sleep(0.5)
            pyautogui.press('tab')
        pyautogui.press('enter')
        time.sleep(2)


if __name__ == "__main__":
    download_path = input("请输入保存文件夹的绝对路径：")
    print(download_path)
    Stanford = Download_Article(download_path)
    Stanford.download()
